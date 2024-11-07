import pandas as pd
import os
from typing import Tuple, List
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side

class PivotProcessor:
    def __init__(self, file_path: str, result_path: str):
        self.file_path = file_path
        self.result_path = result_path

    @staticmethod
    def get_sheets(file_path: str) -> List[str]:
        """获取Excel文件中的所有工作表名称"""
        try:
            with pd.ExcelFile(file_path) as xl:
                return xl.sheet_names
        except Exception as e:
            raise Exception(f"读取工作表失败: {str(e)}")

    @staticmethod
    def get_sheet_info(file_path: str, sheet_name: str) -> dict:
        """获取工作表的列信息"""
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            headers = df.columns.tolist()
            columns = [chr(65 + i) for i in range(len(headers))]
            return {
                'headers': headers,
                'columns': columns
            }
        except Exception as e:
            raise Exception(f"读取列信息失败: {str(e)}")

    def process(self, sheet: str, config: dict) -> Tuple[bool, str]:
        """处理数据透视表操作"""
        try:
            # 读取数据
            df = pd.read_excel(self.file_path, sheet_name=sheet)
            
            # 获取列名
            def get_column_index(col_letter: str) -> int:
                """将Excel列标识符（A、B、C...）转换为数字索引（0、1、2...）"""
                result = 0
                for char in col_letter.upper():
                    result = result * 26 + (ord(char) - ord('A'))
                return result

            # 获取行标签、列标签和值列
            row_cols = [get_column_index(col) for col in config.get('rows', [])]
            col_cols = [get_column_index(col) for col in config.get('cols', [])] if config.get('cols') else []
            value_configs = config.get('values', [])

            # 检查每个值列的数据类型和聚合函数
            for vc in value_configs:
                col_idx = get_column_index(vc['column'])
                col_name = df.columns[col_idx]
                agg_func = vc['aggfunc']
                
                # 检查是否可以转换为数值
                try:
                    pd.to_numeric(df[col_name], errors='raise')
                except:
                    # 如果不能转换为数值，只允许计数操作
                    if agg_func != 'count':
                        return False, f"列 '{col_name}' 是非数值类型，只支持计数操作"

            # 获取列名
            def get_col_name(col_letter: str) -> str:
                """获取列的实际名称"""
                idx = get_column_index(col_letter)
                return df.columns[idx]

            # 创建透视表
            pivot_data = []
            
            # 按行标签和列标签分组
            group_cols = [df.columns[i] for i in (row_cols + col_cols)]
            grouped = df.groupby(group_cols)
            
            # 处理每个分组
            for name, group in grouped:
                row_data = {}
                
                # 处理标签
                if isinstance(name, tuple):
                    for i, col in enumerate(group_cols):
                        row_data[col] = name[i]
                else:
                    row_data[group_cols[0]] = name
                
                # 处理每个值列的计算
                for vc in value_configs:
                    col = get_col_name(vc['column'])
                    agg_func = vc['aggfunc']
                    
                    # 对于非数值列，只进行计数
                    try:
                        pd.to_numeric(df[col], errors='raise')
                        # 数值列可以进行所有聚合操作
                        if agg_func == 'sum':
                            value = group[col].sum()
                        elif agg_func == 'mean':
                            value = group[col].mean()
                        elif agg_func == 'count':
                            value = len(group)
                        elif agg_func == 'max':
                            value = group[col].max()
                        elif agg_func == 'min':
                            value = group[col].min()
                    except:
                        # 非数值列只进行计数
                        value = len(group)
                    
                    row_data[f"{col}({agg_func})"] = value
                
                pivot_data.append(row_data)
            
            # 创建结果DataFrame
            result = pd.DataFrame(pivot_data)
            
            # 添加总计行
            totals = {group_cols[0]: '总计'}
            if len(group_cols) > 1:
                for col in group_cols[1:]:
                    totals[col] = ''
            for col in result.columns:
                if col not in group_cols:
                    totals[col] = result[col].sum()
            
            # 使用concat替代append
            result = pd.concat([result, pd.DataFrame([totals])], ignore_index=True)

            # 格式化数值
            for col in result.columns:
                if col not in group_cols:
                    result[col] = result[col].apply(lambda x: 
                        '{:,.2f}'.format(x) if 'mean' in col 
                        else '{:,}'.format(int(x)))

            # 保存结果
            with pd.ExcelWriter(self.result_path, engine='openpyxl') as writer:
                result.to_excel(writer, index=False)
                
                # 获取工作表
                worksheet = writer.sheets['Sheet1']
                
                # 设置列宽
                for idx, col in enumerate(worksheet.columns, 1):
                    max_length = 0
                    for cell in col:
                        try:
                            max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    worksheet.column_dimensions[get_column_letter(idx)].width = max_length + 2

                # 设置样式
                # 定义边框样式
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                # 合并相同值的单元格
                def merge_same_values(start_row, end_row, col):
                    current_value = None
                    merge_start = None
                    
                    for row in range(start_row, end_row + 1):
                        cell_value = worksheet.cell(row=row, column=col).value
                        
                        if cell_value == current_value:
                            continue
                        else:
                            # 如果有需要合并的单元格
                            if merge_start and merge_start < row - 1:
                                worksheet.merge_cells(
                                    start_row=merge_start,
                                    end_row=row - 1,
                                    start_column=col,
                                    end_column=col
                                )
                                # 设置合并后的单元格样式
                                merged_cell = worksheet.cell(row=merge_start, column=col)
                                merged_cell.alignment = Alignment(horizontal='left', vertical='center')
                            
                            current_value = cell_value
                            merge_start = row
                    
                    # 处理最后一组相同的值
                    if merge_start and merge_start < end_row:
                        worksheet.merge_cells(
                            start_row=merge_start,
                            end_row=end_row,
                            start_column=col,
                            end_column=col
                        )
                        merged_cell = worksheet.cell(row=merge_start, column=col)
                        merged_cell.alignment = Alignment(horizontal='left', vertical='center')

                # 获取数据范围
                max_row = worksheet.max_row
                max_col = worksheet.max_column

                # 对每个标签列进行合并
                for col_idx in range(1, len(group_cols) + 1):
                    merge_same_values(2, max_row - 1, col_idx)  # 从第2行开始，跳过标题行和总计行

                # 设置所有单元格的基本样式
                for row in worksheet.iter_rows(min_row=1, max_row=max_row):
                    for cell in row:
                        cell.border = thin_border
                        if row[0].row == 1:  # 标题行
                            cell.font = Font(bold=True)
                            cell.alignment = Alignment(horizontal='center')
                        elif row[0].row == max_row:  # 总计行
                            cell.font = Font(bold=True)
                        else:  # 数据行
                            if cell.column <= len(group_cols):  # 标签列
                                if not cell.alignment:  # 如果不是合并的单元格
                                    cell.alignment = Alignment(horizontal='left')
                            else:  # 值列
                                cell.alignment = Alignment(horizontal='right')

            return True, ""

        except Exception as e:
            print("错误:", str(e))
            import traceback
            traceback.print_exc()
            return False, f"创建透视表失败: {str(e)}"

    def cleanup(self):
        """清理临时文件"""
        try:
            if os.path.exists(self.file_path):
                os.remove(self.file_path)
        except Exception as e:
            print(f"清理文件时出错: {e}")